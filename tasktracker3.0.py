# -*- coding: utf-8 -*-
"""
Created on Thu Jul  3 21:34:12 2025

@author: malav
"""

import streamlit as st
import pandas as pd
from datetime import datetime
import io
import json

# Initialize session state
if 'tasks' not in st.session_state:
    st.session_state.tasks = []

def export_to_excel(tasks):
    """Export tasks to Excel file"""
    if not tasks:
        return None
    
    df = pd.DataFrame(tasks)
    output = io.BytesIO()
    
    try:
        # Try using xlsxwriter first
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Task_Sheets')
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Task_Sheets']
            
            # Add formatting
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D7E4BD',
                'border': 1
            })
            
            # Write headers with formatting
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                
            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.set_column(i, i, max_length)
        
    except ImportError:
        # Fallback to openpyxl if xlsxwriter is not available
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Task_Sheets')
            
            # Basic formatting with openpyxl
            workbook = writer.book
            worksheet = writer.sheets['Task_Sheets']
            
            # Style the header row
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D7E4BD', end_color='D7E4BD', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, col_letter in enumerate(worksheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(df.columns))):
                for cell in col_letter:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

def export_to_json(tasks):
    """Export tasks to JSON file"""
    if not tasks:
        return None
    
    json_data = json.dumps(tasks, indent=2, default=str)
    return json_data.encode('utf-8')

def export_to_csv(tasks):
    """Export tasks to CSV file"""
    if not tasks:
        return None
    
    df = pd.DataFrame(tasks)
    output = io.StringIO()
    df.to_csv(output, index=False)
    return output.getvalue().encode('utf-8')

def delete_task(task_index):
    """Delete a task by index"""
    if 0 <= task_index < len(st.session_state.tasks):
        st.session_state.tasks.pop(task_index)
        st.success("Task deleted successfully!")
        st.rerun()

# Streamlit App Configuration
st.set_page_config(
    page_title="Task Sheet Tracker",
    page_icon="📋",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# Custom CSS to match your design
st.markdown("""
<style>
    .main-header {
        background-color: #A8C686;
        padding: 20px;
        border-radius: 10px;
        margin-bottom: 30px;
        text-align: center;
    }
    
    .form-container {
        background-color: #A8C686;
        padding: 30px;
        border-radius: 10px;
        margin: 20px 0;
    }
    
    .stTextInput > div > div > input {
        background-color: white;
        border: 2px solid #666;
        border-radius: 5px;
    }
    
    .stDateInput > div > div > input {
        background-color: white;
        border: 2px solid #666;
        border-radius: 5px;
    }
    
    .stTextArea > div > div > textarea {
        background-color: white;
        border: 2px solid #666;
        border-radius: 5px;
    }
    
    .stSelectbox > div > div > select {
        background-color: white;
        border: 2px solid #666;
        border-radius: 5px;
    }
    
    .icon-text {
        display: flex;
        align-items: center;
        font-size: 18px;
        margin-bottom: 10px;
    }
    
    .submit-btn {
        background-color: #4CAF50;
        color: white;
        padding: 10px 20px;
        border: none;
        border-radius: 5px;
        cursor: pointer;
        width: 100%;
        margin-top: 20px;
    }
    
    .metric-container {
        background-color: #f0f2f6;
        padding: 15px;
        border-radius: 10px;
        text-align: center;
        border-left: 4px solid #A8C686;
    }
    
    .task-card {
        background-color: #f8f9fa;
        padding: 15px;
        border-radius: 8px;
        border-left: 4px solid #A8C686;
        margin-bottom: 10px;
    }
</style>
""", unsafe_allow_html=True)

# Main App
def main():
    # Header
    st.markdown('<div class="main-header"><h1>📋 Task Sheet Tracker</h1><p>Local Version - No External Database Required</p></div>', unsafe_allow_html=True)
    
    # Create tabs
    tab1, tab2, tab3, tab4 = st.tabs(["➕ Add Task", "📋 View Tasks", "📊 Export Data", "⚙️ Manage Data"])
    
    with tab1:
        st.markdown('<div class="form-container">', unsafe_allow_html=True)
        
        # Task Entry Form
        with st.form("task_form", clear_on_submit=True):
            col1, col2 = st.columns([1, 4])
            
            with col1:
                st.markdown("👤")
            with col2:
                name = st.text_input("Name", placeholder="Enter your name")
            
            col1, col2 = st.columns([1, 4])
            with col1:
                st.markdown("📅")
            with col2:
                date = st.date_input("Date", value=datetime.now())
            
            col1, col2 = st.columns([1, 4])
            with col1:
                st.markdown("📧")
            with col2:
                email = st.text_input("Email", placeholder="Enter your email")
            
            col1, col2 = st.columns([1, 4])
            with col1:
                st.markdown("🏷️")
            with col2:
                priority = st.selectbox("Priority", ["Low", "Medium", "High", "Urgent"])
            
            col1, col2 = st.columns([1, 4])
            with col1:
                st.markdown("📄")
            with col2:
                task_alloted = st.text_area("Task Alloted", placeholder="Describe the task", height=100)
            
            col1, col2 = st.columns([1, 4])
            with col1:
                st.markdown("📝")
            with col2:
                notes = st.text_area("Additional Notes (Optional)", placeholder="Any additional notes or comments", height=80)
            
            submitted = st.form_submit_button("Submit Task", use_container_width=True)
            
            if submitted:
                if name and email and task_alloted:
                    task_data = {
                        'id': len(st.session_state.tasks) + 1,
                        'name': name,
                        'date': str(date),
                        'email': email,
                        'priority': priority,
                        'task_alloted': task_alloted,
                        'notes': notes,
                        'status': 'Pending',
                        'created_at': datetime.now().isoformat()
                    }
                    
                    st.session_state.tasks.append(task_data)
                    st.success("✅ Task submitted successfully!")
                    st.balloons()
                else:
                    st.error("❌ Please fill in all required fields (Name, Email, Task)")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tab2:
        st.subheader("📋 All Tasks")
        
        tasks = st.session_state.tasks
        
        if tasks:
            df = pd.DataFrame(tasks)
            
            # Display summary
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Tasks", len(tasks))
            with col2:
                unique_users = len(df['name'].unique()) if 'name' in df.columns else 0
                st.metric("Unique Users", unique_users)
            with col3:
                today_tasks = len(df[df['date'] == str(datetime.now().date())]) if 'date' in df.columns else 0
                st.metric("Today's Tasks", today_tasks)
            with col4:
                pending_tasks = len(df[df['status'] == 'Pending']) if 'status' in df.columns else 0
                st.metric("Pending Tasks", pending_tasks)
            
            st.markdown("---")
            
            # Filter options
            col1, col2, col3 = st.columns(3)
            with col1:
                name_filter = st.selectbox("Filter by Name", ["All"] + list(df['name'].unique()))
            with col2:
                priority_filter = st.selectbox("Filter by Priority", ["All", "Low", "Medium", "High", "Urgent"])
            with col3:
                status_filter = st.selectbox("Filter by Status", ["All", "Pending", "In Progress", "Completed"])
            
            # Apply filters
            filtered_df = df.copy()
            if name_filter != "All":
                filtered_df = filtered_df[filtered_df['name'] == name_filter]
            if priority_filter != "All":
                filtered_df = filtered_df[filtered_df['priority'] == priority_filter]
            if status_filter != "All":
                filtered_df = filtered_df[filtered_df['status'] == status_filter]
            
            # Display filtered tasks
            if not filtered_df.empty:
                for i, (idx, task) in enumerate(filtered_df.iterrows()):
                    priority_color = {
                        'Low': '🟢',
                        'Medium': '🟡', 
                        'High': '🟠',
                        'Urgent': '🔴'
                    }.get(task.get('priority', 'Medium'), '🟡')
                    
                    with st.expander(f"Task {task.get('id', i+1)}: {task.get('name', 'Unknown')} - {task.get('date', 'No date')} {priority_color}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write(f"**👤 Name:** {task.get('name', 'N/A')}")
                            st.write(f"**📅 Date:** {task.get('date', 'N/A')}")
                            st.write(f"**🏷️ Priority:** {task.get('priority', 'N/A')} {priority_color}")
                        with col2:
                            st.write(f"**📧 Email:** {task.get('email', 'N/A')}")
                            st.write(f"**📊 Status:** {task.get('status', 'Pending')}")
                        
                        st.write(f"**📄 Task:** {task.get('task_alloted', 'N/A')}")
                        if task.get('notes'):
                            st.write(f"**📝 Notes:** {task.get('notes', 'N/A')}")
                        
                        # Task management buttons
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            if st.button(f"Mark Complete", key=f"complete_{idx}"):
                                st.session_state.tasks[idx]['status'] = 'Completed'
                                st.success("Task marked as completed!")
                                st.rerun()
                        with col2:
                            if st.button(f"In Progress", key=f"progress_{idx}"):
                                st.session_state.tasks[idx]['status'] = 'In Progress'
                                st.success("Task marked as in progress!")
                                st.rerun()
                        with col3:
                            if st.button(f"🗑️ Delete", key=f"delete_{idx}", type="secondary"):
                                if st.session_state.get(f'confirm_delete_{idx}', False):
                                    delete_task(idx)
                                else:
                                    st.session_state[f'confirm_delete_{idx}'] = True
                                    st.warning("Click again to confirm deletion")
            else:
                st.info("No tasks match the selected filters.")
        else:
            st.info("No tasks found. Add some tasks in the 'Add Task' tab!")
    
    with tab3:
        st.subheader("📊 Export Data")
        
        tasks = st.session_state.tasks
        
        if tasks:
            st.write(f"Found {len(tasks)} tasks to export")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # Export to Excel
                try:
                    excel_file = export_to_excel(tasks)
                    
                    if excel_file:
                        st.download_button(
                            label="📥 Download Excel File",
                            data=excel_file,
                            file_name=f"task_sheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"Excel export failed: {str(e)}")
                    st.info("Try CSV export instead")
            
            with col2:
                # Export to CSV (fallback option)
                csv_file = export_to_csv(tasks)
                
                if csv_file:
                    st.download_button(
                        label="📥 Download CSV File",
                        data=csv_file,
                        file_name=f"task_sheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
            
            with col3:
                # Export to JSON
                json_file = export_to_json(tasks)
                
                if json_file:
                    st.download_button(
                        label="📥 Download JSON File",
                        data=json_file,
                        file_name=f"task_sheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                        mime="application/json",
                        use_container_width=True
                    )
            
            # Display preview
            st.subheader("Data Preview")
            df = pd.DataFrame(tasks)
            st.dataframe(df, use_container_width=True)
            
            # Statistics
            st.subheader("📈 Statistics")
            col1, col2 = st.columns(2)
            
            with col1:
                if 'priority' in df.columns:
                    priority_counts = df['priority'].value_counts()
                    st.write("**Tasks by Priority:**")
                    st.bar_chart(priority_counts)
            
            with col2:
                if 'status' in df.columns:
                    status_counts = df['status'].value_counts()
                    st.write("**Tasks by Status:**")
                    st.bar_chart(status_counts)
                    
        else:
            st.info("No data to export. Add some tasks first!")
    
    with tab4:
        st.subheader("⚙️ Data Management")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Import Data**")
            uploaded_file = st.file_uploader("Upload JSON file", type=['json'])
            
            if uploaded_file is not None:
                try:
                    content = uploaded_file.read()
                    imported_tasks = json.loads(content)
                    
                    if st.button("Import Tasks"):
                        st.session_state.tasks.extend(imported_tasks)
                        st.success(f"Successfully imported {len(imported_tasks)} tasks!")
                        st.rerun()
                        
                except Exception as e:
                    st.error(f"Error importing file: {str(e)}")
        
        with col2:
            st.write("**Clear Data**")
            st.warning("This will delete all tasks permanently!")
            
            if st.button("🗑️ Clear All Tasks", type="secondary"):
                if st.session_state.get('confirm_clear', False):
                    st.session_state.tasks = []
                    st.session_state.confirm_clear = False
                    st.success("All tasks cleared!")
                    st.rerun()
                else:
                    st.session_state.confirm_clear = True
                    st.error("Click again to confirm deletion of all tasks")
        
        # Data info
        st.markdown("---")
        st.info(f"""
        **Current Session Info:**
        - Total tasks in memory: {len(st.session_state.tasks)}
        - Data persistence: Session only (data will be lost when browser is closed)
        - Export options: Excel and JSON formats available
        """)

# Info sidebar
with st.sidebar:
    st.header("ℹ️ About This App")
    st.markdown("""
    **Task Sheet Tracker (Local Version)**
    
    This is a local version that doesn't require any external database setup.
    
    **Features:**
    - ✅ Add and manage tasks
    - 📊 Priority and status tracking
    - 🔍 Filter and search tasks
    - 📥 Export to Excel/CSV/JSON
    - 📈 Basic statistics
    - 🗑️ Delete individual tasks
    
    **Data Storage:**
    - Tasks are stored in browser session
    - Data persists during the session
    - Export data before closing browser
    
    **Required packages:**
    ```bash
    pip install streamlit pandas
    # For Excel export (optional):
    pip install openpyxl
    # OR
    pip install xlsxwriter
    ```
    """)
    
    st.markdown("---")
    st.markdown("**💡 Tip:** If Excel export fails, use CSV or JSON export instead!")

if __name__ == "__main__":
    main()